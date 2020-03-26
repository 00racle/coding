import paramiko
import pymysql
import openpyxl
import re
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side


conn_db = pymysql.connect(host = '192.168.6.160', user='chk1417', password='8282op82@#', db='dchk')
curs = conn_db.cursor()

#sql = 'select ip, port, id, AES_DECRYPT(UNHEX(pw), ip) as pw, hostname from v_svr;'
#sql = 'select ip, port, id, AES_DECRYPT(UNHEX(pw), ip) as pw, hostname from v_svr where hostname="sobi-webwas";'
#sql = 'select ip, port, id, AES_DECRYPT(UNHEX(pw), ip) as pw, hostname from v_svr where hostname not in("messenger");'
#sql = 'select ip, port, id, AES_DECRYPT(UNHEX(pw), ip) as pw, hostname from t_svr where hostname="Test6";'
sql = 'select ip, port, id, AES_DECRYPT(UNHEX(pw), ip) as pw, hostname from t_svr where hostname="cent7";'
#sql = 'select ip, port, id, AES_DECRYPT(UNHEX(pw), ip) as pw, hostname from t_svr where version like "%cent%";'
#sql = 'select ip, port, id, AES_DECRYPT(UNHEX(pw), ip) as pw, hostname from v_svr where hostname="mail";'
curs.execute(sql)

result = curs.fetchall()
conn_db.close()

ssh = paramiko.SSHClient()
ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

wb = openpyxl.Workbook()                            # 엑셀 생성
sheet = wb.active

def sec_chk(idx, svr):
    try:
        ssh.connect(svr[0], username=svr[2], port=svr[1], password=svr[3].decode('utf-8'))
        ssh.invoke_shell()
    
        stdin, stdout, stderr = ssh.exec_command("cat -n /etc/system-release | cut -d'.' -f1 | awk '{print $NF}'")
        for line in stdout:
            Lver = line.rstrip('\n')   

        print(svr[4])
        thin_border = Border(left=Side(style='thin'),
                right = Side(style='thin'),
                top = Side(style='thin'),
                bottom = Side(style='thin'))

        nsheet = "sheet"+"idx"
        nsheet = wb.create_sheet(svr[4])
        nsheet.column_dimensions['B'].width = 60
        nsheet.column_dimensions['D'].width = 60
        nsheet.column_dimensions['E'].width = 60
        nsheet.cell(row=1, column=1).value = "구분"
        nsheet['A1'].fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
        nsheet['A1'].font = openpyxl.styles.fonts.Font(bold=True)
        nsheet.cell(row=1, column=2).value = "보안대책"
        nsheet['B1'].fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
        nsheet['B1'].font = openpyxl.styles.fonts.Font(bold=True)
        nsheet.cell(row=1, column=3).value = "이행실태"
        nsheet['C1'].fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
        nsheet['C1'].font = openpyxl.styles.fonts.Font(bold=True)
        nsheet.cell(row=1, column=4).value = "점검 명령어"
        nsheet['D1'].fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
        nsheet['D1'].font = openpyxl.styles.fonts.Font(bold=True)
        nsheet.cell(row=1, column=5).value = "명령어 결과"
        nsheet['E1'].fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
        nsheet['E1'].font = openpyxl.styles.fonts.Font(bold=True)

# ================================================== U-01 ==================================================
    
        U01 = ["U-01", "사용자 계정에 0-99번 사이의 UID, GID값이 할당된 계정 확인", "", "cat /etc/passwd | grep -v nologin | awk -F: '$3 < 99'"]
        stdin, stdout, stderr = ssh.exec_command(U01[3])
        u01 = []
        u01s = ""
        for line in stdout:
            u01.append(line.rstrip('\n')) 
        
        U01[2] = "취약" if len(u01) != 1 else "양호"

        for i in u01:
            u01s += i+"\n"

        U01.append(u01s)
        nsheet.append(U01)
        if U01[2] == "취약":
            nsheet['C2'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# ================================================== U-02 ==================================================
  
        U02 = ["U-02", "root 계정 외에 UID와 GID가 0인 계정이 없는지 검검", "", "cat /etc/passwd | grep ':0'"]
        stdin, stdout, stderr = ssh.exec_command(U02[3])
        u02 = []
        u02s = ""
        for line in stdout:
            u02.append(line.rstrip('\n'))
        
        U02[2] = "양호" if len(u02) == 1 else "취약"

        for i in u02:
            u02s += i+"\n"

        U02.append(u02s)
        nsheet.append(U02)
        if U02[2] == "취약":
            nsheet['C3'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# ================================================== U-03 ==================================================

        U03 = ["U-03", "디폴트 시스템 계정 제거", "", "cat /etc/passwd | awk -F: '$3 < 99'"]
        stdin, stdout, stderr = ssh.exec_command(U03[3])
        u03 = []
        u03s = ""
        for line in stdout:
            u03.append(line.rstrip('\n'))
        
        U03[2] = "양호" if len(u03) == 1 else "취약"

        for i in u03:
            u03s += i+"\n"

        U03.append(u03s)
        nsheet.append(U03)
        if U03[2] == "취약":
            nsheet['C4'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# ================================================== U-04 ==================================================

        U04 = ["U-04", "로그인 실패 횟수 제한", "", "cat /etc/pam.d/password-auth |grep -v '#' |grep pam_tally2.so |grep deny | cut -d'=' -f2 | awk '{print $1}'"]
        stdin, stdout, stderr = ssh.exec_command(U04[3])
        u04 = []
        u04s = ""
        for line in stdout:
            u04.append(line.rstrip('\n'))
        
        U04[2] = "취약" if len(u04) == 0 or int(u04[0]) >= 5 else "양호"

        for i in u04:
            u04s += i+"\n"

        U04.append(u04s)
        nsheet.append(U04)
        if U04[2] == "취약":
            nsheet['C5'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# ================================================== U-05 ==================================================

        U05 = ["U-05", "/etc/hosts 600 권한 설정", "", "ls -al /etc/hosts | awk '{print $1}'"]
        stdin, stdout, stderr = ssh.exec_command(U05[3])
        u05 = []
        for line in stdout:
            u05.append(line.rstrip('\n'))

        if u05[0].find('-rw-------.') >= 0:
            U05[2] = "양호"
            u05s = "/etc/hosts 권한은 600 입니다. 결과값: "
        else:
            U05[2] = "취약"
            u05s = "/etc/hosts 권한은 600이 아닙니다. 결과값: "

        for i in u05:
            u05s += i+"\n"

        U05.append(u05s)
        nsheet.append(U05)
        if U05[2] == "취약":
            nsheet['C6'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# ================================================== U-06 ==================================================

        u06 = []
        U06 = ["U-06", "관리자 계정에 대한 로그인 성공/실패 기록 설정", "", "cat /etc/rsyslog.conf | grep -v '#' | grep 'authpriv.*' | grep '/var/log/secure' | wc -l"]
        if Lver == "6":
            stdin, stdout, stderr = ssh.exec_command("sudo service rsyslog status | grep 'running' | wc -l")
            for line in stdout:
                u06.append(line.rstrip('\n'))
        else:
            stdin, stdout, stderr = ssh.exec_command("sudo systemctl status rsyslog | grep 'Active:' | grep 'running' |wc -l")
            for line in stdout:
                u06.append(line.rstrip('\n'))

        stdin, stdout, stderr = ssh.exec_command(U06[3])
        for line in stdout:
            u06.append(line.rstrip('\n'))

        U06[2] = "양호" if u06[0] == "1" and u06[1] == "1" else "취약"
        if u06[0] == "1" and u06[1] == "1":
            U06[2] = "양호"
            u06s = "rsyslog 서비스가 동작중이고, /etc/rsyslog.conf 설정값이 정상입니다."
        elif u06[0] == "0" and u06[1] == "1":
            U06[2] = "취약"
            u06s = "/etc/rsyslog.conf 설정이 정상이나, rsyslog 서비스가 동작중이지 않습니다."
        else:
            U06[2] = "취약"
            u06s = "rsyslog 서비스가 동작중이지 않고, /etc/rsyslog.conf 설정이 취약합니다."

        U06.append(u06s)
        nsheet.append(U06)
        if U06[2] == "취약":
            nsheet['C7'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# ================================================== U-07 ==================================================

        U07 = ["U-07", "네트워크 서비스 데몬 권한 755 초과 여부 확인", "", "sudo ls -l /etc/xinetd.conf"]
        u07 = []
        u07s = ""
        stdin, stdout, stderr = ssh.exec_command(U07[3])
        for line in stdout:
            u07.append(line.rstrip('\n'))
        for line in stderr:
            u07.append(line.rstrip('\n'))
    
        #if u07[0].find("No such file or directory") != -1:
        if u07[0].find("cannot access") != -1:
            U07[2] = "양호"
            u07s = "xinetd.conf 파일이 없습니다: "
        elif re.search("-..-------.", u07[0]) and re.search("root root", u07[0]):
            U07[2] = "양호"
            u07s= "xinetd.conf 파일의 권한 및 소유자설정 양호: "
        else:
            U07[2] = "취약"
            u07s = "xinetd.conf 파일의 권한 및 소유자설정 취약: "

        for i in u07:
            u07s += i+"\n"

        U07.append(u07s)
        nsheet.append(U07)
        if U07[2] == "취약":
            nsheet['C08'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# ================================================== U-08 ==================================================

        u08 = []
        U08 = ["U-08", "안전한 비밀번호 설정(9자리 이상, 숫자, 영문자, 특수문자 혼용", "", ""]
        if Lver == "6":
            U08[3] = "cat /etc/pam.d/system-auth | grep -v '#' | grep 'pam_cracklib.so' |awk '{print $7, $8, $9, $10}'"
            stdin, stdout, stderr = ssh.exec_command(U08[3])
            for line in stdout:
                u08.append(line.rstrip('\n'))
            u08 = u08[0].split(" ")
        else:
            U08[3] = "cat /etc/security/pwquality.conf | grep -v '#'"
            stdin, stdout, stderr = ssh.exec_command(U08[3])
            for line in stdout:
                u08.append(line.rstrip('\n'))
        cnt = 0
        for i in u08:
            if "minlen" in i and int(i.split('=')[1]) >= 9:
                cnt += 1
            elif "dcredit" in i and abs(int(i.split('=')[1])) >= 0:
                cnt += 1
            elif "lcredit" in i and abs(int(i.split('=')[1])) >= 0:
                cnt += 1
            elif "ocredit" in i and abs(int(i.split('=')[1])) >= 0:
                cnt += 1
        u08s = " ".join(u08)
        if cnt == 4:   
            U08[2] = "양호"
        else:
            U08[2] = "취약"

        U08.append(u08s)
        nsheet.append(U08)
        if U08[2] == "취약":
            nsheet['C9'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# ================================================== U-09 ==================================================

        stdin, stdout, stderr = ssh.exec_command("ls -al /etc/passwd | awk '{print $3, $4}'")
        u09 = []
        u09s = "/etc/passwd = "
        U09 = ["U-09", "패스워드 관리 시스템 파일의 소유자를 슈퍼관리자로 지정여부", "", "ls -al /etc/passwd, /etc/shadow"]
        for line in stdout:
            u09.append(line.rstrip('\n'))
        stdin, stdout, stderr = ssh.exec_command("ls -al /etc/shadow | awk '{print $3, $4}'")
        for line in stdout:
            u09.append(line.rstrip('\n'))
    
        if u09[0] == "root root" and u09[1] == "root root":
            U09[2] = "양호"
        else:
            U09[2] = "취약"
        u09s = u09s+u09[0]+", "+"/etc/shadow = "+u09[1]

        U09.append(u09s)
        nsheet.append(U09)
        if U09[2] == "취약":
            nsheet['C10'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# ================================================== U-10 ==================================================

        stdin, stdout, stderr = ssh.exec_command("ls -al /etc/passwd | awk '{print $1}'")
        u10 = []
        u10s = "/etc/passwd = "
        U10 = ["U-10", "패스워드 관리를 위한 시스템 파일은 슈퍼관리자만 수정 권한 소유 확인", "", "ls -al /etc/passwd, /etc/shadow"]
        for line in stdout:
            u10.append(line.rstrip('\n'))
        stdin, stdout, stderr = ssh.exec_command("ls -al /etc/shadow | awk '{print $1}'")
        for line in stdout:
            u10.append(line.rstrip('\n'))

        if re.search('-..-.--.--', u10[0]) and re.search('-.--------', u10[1]):
            U10[2] = "양호"
        else:
            U10[2] = "취약"
        u10s = u10s+u10[0]+", "+"/etc/shadow = "+u10[1]

        U10.append(u10s)
        nsheet.append(U10)
        if U10[2] == "취약":
            nsheet['C11'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# ================================================== U-11 ==================================================

        U11 = ["U-11", "원격 로그인 또는 원격 쉘 등 사용 설정 유무", "보류"]
        nsheet.append(U11)
        if U11[2] == "취약":
            nsheet['C12'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        elif U11[2] == "보류":
            nsheet['C12'].fill = PatternFill(start_color="99CC33", end_color="99CC33", fill_type="solid")

# ================================================== U-12 ==================================================
    
        U12 = ["U-12", "불필요한 네트워크 서비스 제거", "", "rpm -qa | grep -E 'telnet|ftp|rsh|rlogin|rexec|talk|finger'"]
        u12 = []
        u12s = ""
        stdin, stdout, stderr = ssh.exec_command(U12[3])
        for line in stdout:
            u12.append(line.rstrip('\n'))

        if len(u12) == 0:
            U12[2] = "양호"
            u12s = "불필요 네트워크 서비스(telnet, ftp, rsh, rlogin, rexec, talk, finger)가 없습니다."
        else:
            U12[2] = "취약"
            u12s = "불필요 네트워크 서비스가 있습니다: "

        for i in u12:
            u12s += i+"\n"

        U12.append(u12s)
        nsheet.append(U12)
        if U12[2] == "취약":
            nsheet['C13'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# ================================================== U-13 ==================================================

        U13 = ["U-13", "암호화 기능이 없는 프로토콜(Telnet, Ftp)사용 여부 확인", "", "rpm -qa | grep -E 'telnet|ftp'"]
        stdin, stdout, stderr = ssh.exec_command(U13[3])
        u13 = []
        u13s = ""
        for line in stdout:
            u13.append(line.rstrip('\n'))

        if len(u13) == 0:
            U13[2] = "양호"
            u13s = "Telnet, Ftp 서비스 사용하지 않음"
        else:
            U13[2] = "취약"
            u13s = "Telnet 또는 FTP 서비스가 사용중입니다. 사용중 서비스: "
            for i in u13:
                u13s += i+"\n"

        U13.append(u13s)
        nsheet.append(U13)
        if U13[2] == "취약":
            nsheet['C14'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    
# ================================================== U-14 ==================================================

        U14 = ["U-14", "익명 FTP 사용 여부 확인", "", "cat /etc/passwd | grep ftp"]
        stdin, stdout, stderr = ssh.exec_command(U14[3])
        u14 = []
        u14s = ""
        for line in stdout:
            u14.append(line.rstrip('\n'))

        U14[2] = "양호" if len(u14) == 0 else "취약"
        if len(u14) == 0:
            U14[2] = "양호"
            u14s = "/etc/passwd 파일에 ftp 계정이 존재하지 않습니다."
        else:
            U14[2] = "취약"
            u14s = "/etc/passwd 파일에 ftp 계정이 존재합니다. 결과값: "
        
        for i in u14:
            u14s += i+"\n"

        U14.append(u14s)
        nsheet.append(U14)
        if U14[2] == "취약":
            nsheet['C15'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# ================================================== U-15 ==================================================

        U15 = ["U-15", "자동으로 FTP에 로그인을 허용하는 .netrc 파일 제거 확인", "", "sudo find / -name '.netrc'"]
        stdin, stdout, stderr = ssh.exec_command(U15[3])
        u15 = []
        u15s = ""
        for line in stdout:
            u15.append(line.rstrip('\n'))

        U15[2] = "양호" if len(u15) == 0 else "취약" 
        for i in u15:
            u15s += i+"\n"

        U15.append(u15s)
        nsheet.append(U15)
        if U15[2] == "취약":
            nsheet['C16'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# ================================================== U-16 ==================================================

        U16 = ["U-16", "inetd/xinetd가 불필요한 경우 비활성화", "", "rpm -qa | grep -E 'inetd|xinetd'"]
        stdin, stdout, stderr = ssh.exec_command(U16[3])
        u16 = []
        u16s = ""
        for line in stdout:
            u16.append(line.rstrip('\n'))

        U16[2] = "양호" if len(u16) == 0 else "취약" 
        for i in u16:
            u16s += i+"\n"

        U16.append(u16s)
        nsheet.append(U16)
        if U16[2] == "취약":
            nsheet['C17'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# ================================================== U-17 ==================================================

        U17 = ["U-17", "지정된 IP외의 주소에서 FTP나 TELNET으로 접속한 이력여부 점검 확인", "", "sudo tail -5 /var/log/secure | grep -E 'telnet|ftp'"]
        stdin, stdout, stderr = ssh.exec_command(U17[3])
        u17 = []
        u17s = ""
        for line in stdout:
            u17.append(line.rstrip('\n'))

        if U13[2] == "양호" and len(u17) == 0:
            U17[2] = "양호"
        elif U13[2] != "양호" and len(u17) == 0:
            U17[2] = "양호"
        else:
            U17[2] = "취약"
        for i in u17:
            u17s += i+"\n"
    
        U17.append(u17s)
        nsheet.append(U17)
        if U17[2] == "취약":
            nsheet['C18'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    
# ================================================== U-18 ==================================================

        U18 = ["U-18", "Cron 파일 소유자 및 권한 설정 점검", "", "find /etc/ -type f -name cron.* | xargs ls -l"]
        stdin, stdout, stderr = ssh.exec_command(U18[3])
        u18 = []
        u18s = ""
        for line in stdout:
            u18.append(line.rstrip('\n'))
        
        U18[2] = "양호" if re.search('-...------', u18[0]) and re.search('root.root', u18[0]) else "취약" 
        for i in u18:
            u18s += i+"\n"
    
        U18.append(u18s)
        nsheet.append(U18)
        if U18[2] == "취약":
            nsheet['C19'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# ================================================== U-19 ==================================================

        U19 = ["U-19", "디폴트 SNMP 커뮤니티 스트링(public) 값 변경", "", "sudo cat -n /etc/snmp/snmpd.conf | grep com2sec | egrep 'public|private' | grep -v '#'"]
        u19 = []
        u19s = ""
        if Lver == "6":
            stdin, stdout, stderr = ssh.exec_command("sudo service snmpd status | grep -E 'running|실행하고 있습니다' | wc -l")
            for line in stdout:
                u19.append(line.rstrip('\n'))
        else:
            stdin, stdout, stderr = ssh.exec_command("sudo systemctl status snmpd | grep 'Active:' | grep 'running' | wc -l")
            for line in stdout:
                u19.append(line.rstrip('\n'))
        stdin, stdout, stderr = ssh.exec_command(U19[3])
        for line in stdout:
            u19.append(line.rstrip('\n'))
        if len(u19) == 1 and u19[0] == "1":
            U19[2] = "양호"
            u19s = "snmp 서비스가 사용중이나 'public' 커뮤니티값을 사용중이지 않습니다"
        elif len(u19) != 1 and u19[0] == "1":
            U19[2] = "취약"
            u19s = "snmp 서비스가 사용중이고 'public' 커뮤티니값을 사용중입니다. 값: "
            u19s+" "+u19[1]
        else:
            U19[2] = "양호"
            u19s = "snmp 서비스가 사용중이지 않습니다"


        U19.append(u19s) 
        nsheet.append(U19)
        if U19[2] == "취약":
            nsheet['C20'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# ================================================== U-20 ==================================================

        U20 = ["U-20", "최신 버전의 SSH사용", "", "rpm -qa | grep 'openssh'"]
        u20 = []
        cnt = 0
        stdin, stdout, stderr = ssh.exec_command(U20[3])
        for line in stdout:
            u20.append(line.rstrip('\n'))

        if Lver == "6":
            for i in u20:
                cnt += i.find("5.3p1-124.el6_10.x86_64")
            u20s = "openssh 최신 버젼은 [5.3p1-124.el6_10.x86_64], 현재 버전은: "+u20[0][-23:]
        else:
            for i in u20:
                cnt += i.find("7.4p1-21.el7.x86_64")
            u20s = "openssh 최신 버젼은 [7.4p1-21.el7.x86_64], 현재 버전은: "+u20[0][-19:]

        U20[2] = "양호" if cnt == 39 else "취약"
    
        U20.append(u20s)
        nsheet.append(U20)
        if U20[2] == "취약":
            nsheet['C21'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# ================================================== U-21 ==================================================

        #U21 = ["U-21", "IP 포워딩 비활성화", "", "sudo sysctl net.ipv4.ip_forward"]
        U21 = ["U-21", "IP 포워딩 비활성화", "", "/sbin/sysctl net.ipv4.ip_forward"]
        u21 = []
        u21s = ""
        stdin, stdout, stderr = ssh.exec_command(U21[3])
        for line in stdout:
            u21.append(line.rstrip('\n'))
        for line in stderr:
             u21.append(line.rstrip('\n'))

        U21[2] = "양호" if u21[0][-1] == "0" else "취약"
    
        for i in u21:
            u21s += i+"\n"
    
        U21.append(u21s)
        nsheet.append(U21)
        if U21[2] == "취약":
            nsheet['C22'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    
# ================================================== U-22 ==================================================
        U22 = ["U-22", "패킷 재전송 설정 비활성화", "보류", ""]
        nsheet.append(U22)
        if U22[2] == "취약":
            nsheet['C23'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        elif U22[2] == "보류":
            nsheet['C23'].fill = PatternFill(start_color="99CC33", end_color="99CC33", fill_type="solid")
# ================================================== E-01 ==================================================
        E01 = ["E-01", "SSH서비스 기본 포트 변경", ""]

        if svr[1] != 22:
            E01[2] = "양호"
        else:
            E01[2] = "취약"
        
        nsheet.append(E01)
        if E01[2] == "취약":
            nsheet['C24'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# =================================================== END ==================================================



        ssh.close()
     
        for i in range(1, 6):
            for j in range(1, 26):
                nsheet.cell(row=j, column=i).border = thin_border
    except paramiko.ssh_exception.NoValidConnectionsError as 접속불가:
    #except paramiko.SSHException.NoValidConnectionsError as 접속불가:
        print("접속불가(ip및 포트확인)")
    
    except paramiko.ssh_exception.AuthenticationException as 접속불가:
        print("접속불가(인증에러)")

for idx, svr in enumerate(result):
    sec_chk(idx, svr)

#wb.save(filename="Unix_sec_chk.xlsx")
wb.save(filename="chk_cent7.xlsx")
